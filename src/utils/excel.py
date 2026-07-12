import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from utils.db_helpers import fetch_all, fetch_one
from utils.date_helpers import MONTHS_SHORT


C_DARK   = "1A1A2E"
C_MID    = "16213E"
C_ACCENT = "E94560"
C_LIGHT  = "EAF0FB"
C_WHITE  = "FFFFFF"
C_GRAY   = "F2F2F2"

def _hdr(cell, text, bg=C_DARK, bold=True, size=11):
    cell.value = text
    cell.font = Font(bold=bold, size=size, color=C_WHITE)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def build_dashboard_sheet(wb: Workbook, year: int, profile: dict):
    ws = wb.create_sheet("Dashboard")
    ws.merge_cells("A1:D1")
    _hdr(ws["A1"], f"MND DASHBOARD {year}", size=14)
    ws.row_dimensions[1].height = 30

    fields = [
        ("ชื่อ", profile.get("name", "")),
        ("ปีเกิด", profile.get("birth_year", "")),
        ("รายได้ก่อนภาษี/ปี (฿)", profile.get("annual_income", "")),
    ]
    for i, (label, val) in enumerate(fields, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22


def build_monthly_sheet(wb: Workbook, year: int):
    ws = wb.create_sheet("Monthly_Log")
    headers = ["เดือน", "รายได้ (฿)", "รายจ่าย (฿)", "Savings Rate (%)", "Net Worth (฿)", "หมวดบานปลาย", "Note"]
    for i, h in enumerate(headers):
        c = ws.cell(row=1, column=i+1, value=h)
        c.font = Font(bold=True, color=C_WHITE)
        c.fill = PatternFill("solid", fgColor=C_ACCENT)
        c.alignment = Alignment(horizontal="center")

    rows = fetch_all("SELECT * FROM monthly_log WHERE year=? ORDER BY month", (year,))
    row_map = {r["month"]: r for r in rows}
    for month in range(1, 13):
        r = row_map.get(month, {})
        row = [
            MONTHS_SHORT[month - 1],
            r.get("income", ""),
            r.get("expenses", ""),
            r.get("savings_rate", ""),
            r.get("net_worth", ""),
            r.get("top_overspend", ""),
            r.get("note", ""),
        ]
        for j, val in enumerate(row):
            ws.cell(row=month + 1, column=j+1, value=val).border = _border()

    for col in ["A","B","C","D","E","F","G"]:
        ws.column_dimensions[col].width = 18


def build_budget_sheet(wb: Workbook, year: int):
    ws = wb.create_sheet("Annual_Budget")
    headers = ["หมวดหลัก", "Sub-category", "Budget/เดือน (฿)", "Budget/ปี (฿)", "Actual YTD (฿)", "Variance (฿)", "% ใช้ไป"]
    for i, h in enumerate(headers):
        c = ws.cell(row=1, column=i+1, value=h)
        c.font = Font(bold=True, color=C_WHITE)
        c.fill = PatternFill("solid", fgColor=C_ACCENT)
        c.alignment = Alignment(horizontal="center")

    rows = fetch_all("SELECT * FROM budget_categories WHERE year=? ORDER BY category, sub_category", (year,))
    for i, r in enumerate(rows, start=2):
        annual = (r.get("monthly_budget") or 0) * 12
        actual = r.get("actual_ytd") or 0
        variance = annual - actual
        pct = f"{round(actual/annual*100)}%" if annual else ""
        data = [r["category"], r["sub_category"], r.get("monthly_budget",""), annual, actual, variance, pct]
        for j, val in enumerate(data):
            ws.cell(row=i, column=j+1, value=val).border = _border()

    for col, w in zip("ABCDEFG", [18,24,16,16,16,16,10]):
        ws.column_dimensions[col].width = w


def export_workbook(year: int) -> io.BytesIO:
    profile = fetch_one("SELECT * FROM profile WHERE id=1")
    if not profile:
        profile = {}

    wb = Workbook()
    wb.remove(wb.active)
    build_dashboard_sheet(wb, year, profile)
    build_monthly_sheet(wb, year)
    build_budget_sheet(wb, year)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
