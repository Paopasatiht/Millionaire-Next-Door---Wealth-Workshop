"""
Run once to generate templates/wealth-tracker-template.xlsx
Usage: python scripts/create_template.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart

# ── Colors ───────────────────────────────────────────────────────────────────
C_DARK   = "1A1A2E"   # dark navy header
C_MID    = "16213E"   # mid navy
C_ACCENT = "E94560"   # red accent
C_GOLD   = "F5A623"   # gold for highlights
C_LIGHT  = "EAF0FB"   # light blue row
C_WHITE  = "FFFFFF"
C_GRAY   = "F2F2F2"
C_GREEN  = "27AE60"
C_RED_L  = "FADBD8"

def style(ws):
    """Return helper that applies header style to a cell."""
    def h(cell, text, bold=True, size=11, color=C_DARK, bg=None, align="center"):
        cell.value = text
        cell.font = Font(bold=bold, size=size, color=C_WHITE if bg and bg != C_WHITE else color)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
    return h

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def freeze(ws, cell="B2"):
    ws.freeze_panes = cell

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: Dashboard
# ─────────────────────────────────────────────────────────────────────────────
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    h = style(ws)

    # Title
    ws.merge_cells("A1:G1")
    h(ws["A1"], "MND WEALTH DASHBOARD", size=16, bg=C_DARK)
    ws.row_dimensions[1].height = 36

    # Section: Identity
    ws.merge_cells("A3:G3")
    h(ws["A3"], "PROFILE", bg=C_MID)
    labels = ["Name", "Birth Year", "Annual Pre-tax Income (฿)", "Current Year"]
    for i, lbl in enumerate(labels, start=4):
        ws[f"A{i}"] = lbl
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"].fill = PatternFill("solid", fgColor=C_LIGHT)
        ws[f"B{i}"].border = thin_border()

    # Section: PAW/UAW Calculator
    ws.merge_cells("A9:G9")
    h(ws["A9"], "PAW / UAW CALCULATOR", bg=C_MID)
    ws["A10"] = "Age"
    ws["B10"] = "=C7-(YEAR(TODAY())-B4)"   # auto from birth year & income
    ws["A11"] = "Expected Net Worth (฿)"
    ws["B11"] = "=B10*C4/10"
    ws["B11"].number_format = '#,##0'
    ws["A12"] = "Actual Net Worth (฿)"
    ws["B12"].fill = PatternFill("solid", fgColor=C_LIGHT)
    ws["B12"].border = thin_border()
    ws["B12"].number_format = '#,##0'
    ws["A13"] = "Multiplier (Actual / Expected)"
    ws["B13"] = '=IFERROR(B12/B11,"")'
    ws["B13"].number_format = "0.00×"
    ws["A14"] = "Status"
    ws["B14"] = '=IFERROR(IF(B13>=2,"PAW ✓",IF(B13>=0.5,"AAW","UAW ✗")),"")'
    ws["B14"].font = Font(bold=True, size=12)

    # Section: KPIs
    ws.merge_cells("A16:G16")
    h(ws["A16"], "ANNUAL KPIs", bg=C_MID)
    kpi_headers = ["Metric", "Target", "Actual", "Status"]
    for i, hdr in enumerate(kpi_headers):
        c = ws.cell(row=17, column=i+1, value=hdr)
        c.font = Font(bold=True, color=C_WHITE)
        c.fill = PatternFill("solid", fgColor=C_ACCENT)
        c.alignment = Alignment(horizontal="center")

    kpis = [
        ("Savings Rate", "≥ 20%", "=Annual_Budget!C3", ""),
        ("Planning Hours / Month", "≥ 8 hrs", "=AVERAGE(Planning_Time_Log!C2:C13)", ""),
        ("Goals Completed", "100%", "", ""),
        ("Net Worth Growth", "≥ 10%", "", ""),
    ]
    for r, (metric, target, formula, _) in enumerate(kpis, start=18):
        ws.cell(row=r, column=1, value=metric).font = Font(bold=True)
        ws.cell(row=r, column=2, value=target).alignment = Alignment(horizontal="center")
        if formula:
            ws.cell(row=r, column=3, value=formula)
        ws.cell(row=r, column=3).number_format = '0.0%'
        ws.row_dimensions[r].height = 20

    set_col_widths(ws, {"A": 30, "B": 22, "C": 18, "D": 14})
    freeze(ws, "A2")

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: Annual Budget
# ─────────────────────────────────────────────────────────────────────────────
def build_annual_budget(wb):
    ws = wb.create_sheet("Annual_Budget")
    ws.sheet_view.showGridLines = False
    h = style(ws)

    ws.merge_cells("A1:E1")
    h(ws["A1"], "ANNUAL BUDGET", size=14, bg=C_DARK)
    ws.row_dimensions[1].height = 30

    # Summary row
    ws["A2"] = "Total Income (฿)"
    ws["B2"].fill = PatternFill("solid", fgColor=C_LIGHT)
    ws["B2"].border = thin_border()
    ws["A3"] = "Savings Rate"
    ws["B3"] = '=IFERROR(1-(SUM(D5:D100)/B2),"")'
    ws["B3"].number_format = "0.0%"

    # Headers
    headers = ["Category", "Sub-category", "Monthly Budget (฿)", "Annual Budget (฿)", "Actual YTD (฿)", "Variance (฿)", "% Used"]
    for i, hdr in enumerate(headers):
        c = ws.cell(row=5, column=i+1, value=hdr)
        c.font = Font(bold=True, color=C_WHITE)
        c.fill = PatternFill("solid", fgColor=C_ACCENT)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[5].height = 30

    categories = [
        ("อาหาร", "อาหารที่บ้าน"),
        ("", "อาหารนอกบ้าน"),
        ("", "เครื่องดื่ม / กาแฟ"),
        ("ที่อยู่อาศัย", "ค่าเช่า / ค่าผ่อนบ้าน"),
        ("", "ค่าน้ำ / ไฟ / อินเตอร์เน็ต"),
        ("", "ค่าซ่อมแซม / ของใช้ในบ้าน"),
        ("เสื้อผ้า", "เสื้อผ้าตัวเอง"),
        ("", "เสื้อผ้าบุตร"),
        ("การเดินทาง", "น้ำมัน / ค่าเดินทาง"),
        ("", "ค่าบำรุงรักษารถ"),
        ("สุขภาพ", "ประกันสุขภาพ"),
        ("", "ค่าหมอ / ยา"),
        ("การศึกษา", "ค่าเล่าเรียนบุตร"),
        ("", "ค่าหนังสือ / คอร์สเรียน"),
        ("ของขวัญ / สังคม", "ของขวัญ / งานสังคม"),
        ("", "เลี้ยงดูครอบครัว / พ่อแม่"),
        ("บันเทิง", "ท่องเที่ยว"),
        ("", "ความบันเทิงอื่น ๆ"),
        ("ออมทรัพย์ / ลงทุน", "กองทุน / หุ้น"),
        ("", "กองทุนฉุกเฉิน"),
        ("อื่น ๆ", "ค่าใช้จ่ายเบ็ดเตล็ด"),
    ]

    for r, (cat, sub) in enumerate(categories, start=6):
        fill_color = C_LIGHT if r % 2 == 0 else C_WHITE
        ws.cell(row=r, column=1, value=cat).font = Font(bold=bool(cat))
        ws.cell(row=r, column=2, value=sub)
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = thin_border()
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=fill_color)
        # Annual = Monthly * 12
        ws.cell(row=r, column=4).value = f"=C{r}*12"
        ws.cell(row=r, column=4).number_format = "#,##0"
        ws.cell(row=r, column=3).number_format = "#,##0"
        ws.cell(row=r, column=5).number_format = "#,##0"
        # Variance = Annual - Actual
        ws.cell(row=r, column=6).value = f"=D{r}-E{r}"
        ws.cell(row=r, column=6).number_format = "#,##0"
        # % Used
        ws.cell(row=r, column=7).value = f'=IFERROR(E{r}/D{r},"")'
        ws.cell(row=r, column=7).number_format = "0%"

    set_col_widths(ws, {"A": 18, "B": 24, "C": 18, "D": 18, "E": 16, "F": 16, "G": 10})
    freeze(ws, "C6")

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3: Monthly Log
# ─────────────────────────────────────────────────────────────────────────────
def build_monthly_log(wb):
    ws = wb.create_sheet("Monthly_Log")
    ws.sheet_view.showGridLines = False
    h = style(ws)

    ws.merge_cells("A1:N1")
    h(ws["A1"], "MONTHLY LOG", size=14, bg=C_DARK)
    ws.row_dimensions[1].height = 30

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "TOTAL"]
    ws["A2"] = "Metric"
    ws["A2"].font = Font(bold=True, color=C_WHITE)
    ws["A2"].fill = PatternFill("solid", fgColor=C_ACCENT)
    for i, m in enumerate(months):
        c = ws.cell(row=2, column=i+2, value=m)
        c.font = Font(bold=True, color=C_WHITE)
        c.fill = PatternFill("solid", fgColor=C_MID if m != "TOTAL" else C_DARK)
        c.alignment = Alignment(horizontal="center")

    metrics = [
        "รายได้รวม (฿)",
        "รายจ่ายรวม (฿)",
        "ออมได้ (฿)",
        "Savings Rate (%)",
        "Net Worth ณ สิ้นเดือน (฿)",
        "ชั่วโมงวางแผนการเงิน",
        "Budget: ทำตามได้? (Y/N)",
        "หมวดที่บานปลายมากสุด",
        "Note / Lesson",
    ]
    for r, metric in enumerate(metrics, start=3):
        fill_color = C_LIGHT if r % 2 != 0 else C_WHITE
        ws.cell(row=r, column=1, value=metric).font = Font(bold=True)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=C_GRAY)
        for col in range(2, 15):
            cell = ws.cell(row=r, column=col)
            cell.border = thin_border()
            cell.fill = PatternFill("solid", fgColor=fill_color)
            if metric.endswith("(฿)") and col <= 13:
                cell.number_format = "#,##0"
            if metric.endswith("(%)") and col <= 13:
                cell.number_format = "0.0%"
        # TOTAL column formula for numeric rows
        if "฿" in metric or "ชั่วโมง" in metric:
            ws.cell(row=r, column=14).value = f"=SUM(B{r}:M{r})"
        ws.row_dimensions[r].height = 22

    set_col_widths(ws, {"A": 28, **{get_column_letter(i): 10 for i in range(2, 14)}, "N": 12})
    freeze(ws, "B3")

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4: Goals Matrix
# ─────────────────────────────────────────────────────────────────────────────
def build_goals_matrix(wb):
    ws = wb.create_sheet("Goals_Matrix")
    ws.sheet_view.showGridLines = False
    h = style(ws)

    ws.merge_cells("A1:F1")
    h(ws["A1"], "GOALS MATRIX", size=14, bg=C_DARK)
    ws.row_dimensions[1].height = 30

    horizons = [
        ("DAILY HABITS", C_MID, [
            "ตรวจสอบรายจ่ายประจำวัน",
            "อ่านหนังสือ / บทความการเงิน 15 นาที",
            "Custom habit 1",
            "Custom habit 2",
        ]),
        ("WEEKLY GOALS", C_MID, [
            "สรุปรายจ่ายรายสัปดาห์",
            "ทบทวนพอร์ตลงทุน",
            "Custom goal 1",
        ]),
        ("MONTHLY GOALS", C_MID, [
            "กรอก Monthly Log",
            "บันทึกชั่วโมงวางแผนการเงิน",
            "ทบทวน budget vs actual",
            "Custom goal 1",
        ]),
        ("ANNUAL GOALS", C_MID, [
            "เพิ่ม net worth X%",
            "ออมเงิน X฿",
            "Custom goal 1",
            "Custom goal 2",
        ]),
        ("LIFETIME GOALS", C_DARK, [
            "เป้าหมาย net worth สุดท้าย",
            "อายุเกษียณเป้าหมาย",
            "มรดกที่ต้องการทิ้งไว้",
            "Custom goal 1",
        ]),
    ]

    col_headers = ["Goal", "Target / Detail", "Deadline", "Status", "Notes"]
    row = 2
    for horizon, color, goals in horizons:
        ws.merge_cells(f"A{row}:F{row}")
        h(ws[f"A{row}"], horizon, bg=color)
        ws.row_dimensions[row].height = 24
        row += 1

        for i, hdr in enumerate(col_headers):
            c = ws.cell(row=row, column=i+1, value=hdr)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor=C_ACCENT)
            c.font = Font(bold=True, color=C_WHITE)
            c.alignment = Alignment(horizontal="center")
        row += 1

        for goal in goals:
            ws.cell(row=row, column=1, value=goal)
            fill_color = C_LIGHT if row % 2 == 0 else C_WHITE
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border()
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=fill_color)
            ws.row_dimensions[row].height = 20
            row += 1
        row += 1

    set_col_widths(ws, {"A": 30, "B": 28, "C": 14, "D": 12, "E": 22})

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 5: Planning Time Log
# ─────────────────────────────────────────────────────────────────────────────
def build_planning_time_log(wb):
    ws = wb.create_sheet("Planning_Time_Log")
    ws.sheet_view.showGridLines = False
    h = style(ws)

    ws.merge_cells("A1:F1")
    h(ws["A1"], "PLANNING TIME LOG  (target: ≥ 8 hrs/month)", size=13, bg=C_DARK)
    ws.row_dimensions[1].height = 30

    headers = ["Month", "Activities", "Hours", "Topics Studied", "Key Insight", "Source (Book/Podcast/etc.)"]
    for i, hdr in enumerate(headers):
        c = ws.cell(row=2, column=i+1, value=hdr)
        c.font = Font(bold=True, color=C_WHITE)
        c.fill = PatternFill("solid", fgColor=C_ACCENT)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for r, month in enumerate(months, start=3):
        fill_color = C_LIGHT if r % 2 != 0 else C_WHITE
        ws.cell(row=r, column=1, value=month).font = Font(bold=True)
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.border = thin_border()
            cell.fill = PatternFill("solid", fgColor=fill_color)
        ws.cell(row=r, column=3).number_format = "0.0"
        ws.row_dimensions[r].height = 22

    # Total row
    ws.cell(row=15, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=15, column=3, value="=SUM(C3:C14)").font = Font(bold=True)
    ws.cell(row=15, column=3).number_format = "0.0"
    ws.cell(row=15, column=3).fill = PatternFill("solid", fgColor=C_GOLD)

    ws.cell(row=16, column=1, value="Monthly Average").font = Font(bold=True)
    ws.cell(row=16, column=3, value="=AVERAGE(C3:C14)").font = Font(bold=True)
    ws.cell(row=16, column=3).number_format = "0.0"

    ws.cell(row=17, column=1, value="Hit 8hr target? (months)").font = Font(bold=True)
    ws.cell(row=17, column=3, value='=COUNTIF(C3:C14,">=8")').font = Font(bold=True)

    set_col_widths(ws, {"A": 10, "B": 28, "C": 10, "D": 22, "E": 28, "F": 22})
    freeze(ws, "B3")

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 6: Review Log
# ─────────────────────────────────────────────────────────────────────────────
def build_review_log(wb):
    ws = wb.create_sheet("Review_Log")
    ws.sheet_view.showGridLines = False
    h = style(ws)

    ws.merge_cells("A1:D1")
    h(ws["A1"], "ANNUAL REVIEW LOG", size=14, bg=C_DARK)
    ws.row_dimensions[1].height = 30

    sections = [
        ("WEALTH STATUS", [
            ("PAW / UAW Status this year?", ""),
            ("Net Worth vs. last year?", ""),
            ("Savings Rate this year?", ""),
            ("Biggest financial win?", ""),
            ("Biggest financial mistake?", ""),
        ]),
        ("BUDGET REVIEW", [
            ("Did I operate on a written budget?", "Y / N"),
            ("Top 3 categories that overspent?", ""),
            ("Top 3 categories I cut successfully?", ""),
        ]),
        ("GOALS REVIEW", [
            ("Annual goals achieved? (list)", ""),
            ("Goals carried to next year?", ""),
            ("New lifetime goal insight?", ""),
        ]),
        ("PLANNING REVIEW", [
            ("Total planning hours this year?", ""),
            ("Average hours/month?", ""),
            ("Best source of financial education?", ""),
            ("Top 3 lessons learned?", ""),
        ]),
        ("NEXT YEAR COMMITMENTS", [
            ("Annual savings target (฿)?", ""),
            ("Net worth target (฿)?", ""),
            ("One habit to start?", ""),
            ("One habit to stop?", ""),
        ]),
    ]

    row = 3
    for section, items in sections:
        ws.merge_cells(f"A{row}:D{row}")
        h(ws[f"A{row}"], section, bg=C_MID)
        ws.row_dimensions[row].height = 24
        row += 1

        for question, hint in items:
            fill_color = C_LIGHT if row % 2 == 0 else C_WHITE
            ws.cell(row=row, column=1, value=question).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=fill_color)
            ws.merge_cells(f"B{row}:D{row}")
            ws.cell(row=row, column=2, value=hint if hint else "").fill = PatternFill("solid", fgColor=fill_color)
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = thin_border()
            ws.row_dimensions[row].height = 22
            row += 1
        row += 1

    set_col_widths(ws, {"A": 38, "B": 20, "C": 20, "D": 20})

# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_dashboard(wb)
    build_annual_budget(wb)
    build_monthly_log(wb)
    build_goals_matrix(wb)
    build_planning_time_log(wb)
    build_review_log(wb)

    out = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "templates", "wealth-tracker-template.xlsx"
    )
    wb.save(out)
    print(f"Template saved: {out}")

if __name__ == "__main__":
    main()
