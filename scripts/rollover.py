"""
Generate next year's wealth tracker from the current year's file.
Carries forward net worth closing balance as the new year's baseline.

Usage:
    python scripts/rollover.py --source instances/2026/wealth-tracker-2026.xlsx --year 2027
"""

import argparse
import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


def find_net_worth(wb) -> float | None:
    """Read December net worth from Monthly_Log sheet."""
    try:
        ws = wb["Monthly_Log"]
        # Row 7 = Net Worth ณ สิ้นเดือน, column 13 = December (col B=Jan … M=Dec)
        for row in ws.iter_rows():
            if row[0].value and "Net Worth" in str(row[0].value):
                dec_cell = row[12]  # column M = December
                if dec_cell.value is not None:
                    return float(dec_cell.value)
    except Exception as e:
        print(f"Warning: could not read net worth — {e}")
    return None


def set_baseline_net_worth(wb, value: float, year: int):
    """Write closing net worth into Dashboard as the new year's opening baseline."""
    try:
        ws = wb["Dashboard"]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "Actual Net Worth" in str(cell.value):
                    # Write the value into the cell to the right
                    target = ws.cell(row=cell.row, column=cell.column + 1)
                    target.value = value
                    target.number_format = "#,##0"
                    target.fill = PatternFill("solid", fgColor="FFF9C4")
                    print(f"  ✓ Baseline net worth set to ฿{value:,.0f} in Dashboard")
                    return
    except Exception as e:
        print(f"Warning: could not write baseline — {e}")


def update_year_references(wb, old_year: int, new_year: int):
    """Replace year references in sheet names and title cells."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and str(old_year) in cell.value:
                    cell.value = cell.value.replace(str(old_year), str(new_year))


def clear_data_cells(wb):
    """Clear all user-entered data cells, keeping formulas and structure."""
    CLEAR_SHEETS = {
        "Monthly_Log": range(3, 12),       # rows 3–11
        "Planning_Time_Log": range(3, 15), # rows 3–14
        "Annual_Budget": range(6, 28),     # rows 6–27 (actual YTD column only)
        "Review_Log": None,                # clear answer cells
    }

    # Monthly_Log: clear all data columns (B–M)
    if "Monthly_Log" in [ws.title for ws in wb.worksheets]:
        ws = wb["Monthly_Log"]
        for row in CLEAR_SHEETS["Monthly_Log"]:
            for col in range(2, 14):
                ws.cell(row=row, column=col).value = None

    # Planning_Time_Log: clear columns B–F
    if "Planning_Time_Log" in [ws.title for ws in wb.worksheets]:
        ws = wb["Planning_Time_Log"]
        for row in CLEAR_SHEETS["Planning_Time_Log"]:
            for col in range(2, 7):
                ws.cell(row=row, column=col).value = None

    # Annual_Budget: clear Actual YTD column only (column E)
    if "Annual_Budget" in [ws.title for ws in wb.worksheets]:
        ws = wb["Annual_Budget"]
        for row in CLEAR_SHEETS["Annual_Budget"]:
            ws.cell(row=row, column=5).value = None

    # Review_Log: clear answer cells (columns B–D)
    if "Review_Log" in [ws.title for ws in wb.worksheets]:
        ws = wb["Review_Log"]
        for row in ws.iter_rows(min_row=4):
            for cell in row:
                if cell.column > 1 and not cell.font.bold:
                    if isinstance(cell.value, str) and cell.value not in ("Y / N",):
                        cell.value = None

    print("  ✓ Data cells cleared (formulas and structure preserved)")


def main():
    parser = argparse.ArgumentParser(description="MND Wealth Tracker Rollover")
    parser.add_argument("--source", required=True, help="Path to current year's .xlsx file")
    parser.add_argument("--year", required=True, type=int, help="New year (e.g. 2027)")
    args = parser.parse_args()

    if not os.path.exists(args.source):
        print(f"Error: source file not found → {args.source}")
        return

    old_year = args.year - 1
    new_year = args.year

    # Destination path
    dest_dir = os.path.join(os.path.dirname(os.path.dirname(args.source)), str(new_year))
    os.makedirs(dest_dir, exist_ok=True)
    dest = os.path.join(dest_dir, f"wealth-tracker-{new_year}.xlsx")

    print(f"\nMND Rollover: {old_year} → {new_year}")
    print(f"  Source : {args.source}")
    print(f"  Dest   : {dest}")
    print()

    # Load source
    wb_src = load_workbook(args.source)
    net_worth = find_net_worth(wb_src)
    if net_worth:
        print(f"  Found closing net worth: ฿{net_worth:,.0f}")
    else:
        print("  Could not find closing net worth — baseline will be empty")

    # Copy to new file
    shutil.copy2(args.source, dest)
    wb_new = load_workbook(dest)

    # Apply rollover transformations
    clear_data_cells(wb_new)
    if net_worth is not None:
        set_baseline_net_worth(wb_new, net_worth, new_year)
    update_year_references(wb_new, old_year, new_year)

    wb_new.save(dest)
    print(f"\nDone! New file → {dest}")
    print("Open it and update your Profile section in Dashboard.\n")


if __name__ == "__main__":
    main()
